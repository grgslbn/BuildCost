# -*- coding: utf-8 -*-
"""Dump non-empty cells of a sheet as rows. Usage: cv-excel-cells.py <file> <sheet>"""
import openpyxl, sys, os

DIR = r"C:\Users\tieme\Mijn Drive\M²Value\connect value\Excel achter connect Value"
fname = sys.argv[1]
sheet = sys.argv[2] if len(sys.argv) > 2 else None
data_only = "--formula" not in sys.argv

wb = openpyxl.load_workbook(os.path.join(DIR, fname), data_only=data_only, read_only=True)
sheets = [sheet] if sheet else wb.sheetnames
for sname in sheets:
    ws = wb[sname]
    print(f"##### SHEET {sname!r} #####")
    for row in ws.iter_rows():
        cells = [(c.coordinate, c.value) for c in row if c.value not in (None, "")]
        if not cells:
            continue
        parts = []
        for coord, val in cells:
            if isinstance(val, float):
                val = round(val, 2)
            parts.append(f"{coord}={val}")
        print(" | ".join(parts))
wb.close()
