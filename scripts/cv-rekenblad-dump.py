# -*- coding: utf-8 -*-
"""Dump the 170219 rekenbladen.xlsx — the master formula sheet."""
import openpyxl, os

DIR = r"C:\Users\tieme\Mijn Drive\M²Value\connect value\Excel achter connect Value"
wb = openpyxl.load_workbook(os.path.join(DIR, "170219 rekenbladen.xlsx"),
                            data_only=True, read_only=True)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n{'='*90}")
    print(f"SHEET: {sname!r}  rows={ws.max_row}  cols={ws.max_column}")
    print("=" * 90)
    for row in ws.iter_rows(values_only=False):
        cells = []
        for c in row:
            if c.value in (None, ""):
                continue
            v = c.value
            if isinstance(v, float) and v == int(v) and abs(v) < 1e10:
                v = int(v)
            elif isinstance(v, float):
                v = round(v, 4)
            s = str(v)
            if len(s) > 80:
                s = s[:80] + "..."
            cells.append(f"{c.coordinate}={s}")
        if cells:
            print(" | ".join(cells))
wb.close()
