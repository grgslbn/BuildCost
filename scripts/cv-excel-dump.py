# -*- coding: utf-8 -*-
"""Dump structure + cells of the Connect Value source spreadsheets."""
import openpyxl, sys, os

DIR = r"C:\Users\tieme\Mijn Drive\M²Value\connect value\Excel achter connect Value"

files = [
    "connect value bewerkt wat is de m2-prijs.xlsx",
    "170219 rekenbladen.xlsx",
]

def show(path):
    name = os.path.basename(path)
    print("=" * 90)
    print(name)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print("  ERROR:", e)
        return
    for ws in wb.worksheets:
        print(f"  SHEET: {ws.title!r}  max_row={ws.max_row} max_col={ws.max_column}")
    wb.close()

for f in files:
    show(os.path.join(DIR, f))
